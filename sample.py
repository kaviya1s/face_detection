
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QPlainTextEdit
from PyQt5.QtCore import  QObject, pyqtSignal
from PyQt5.QtWidgets import QApplication,QPlainTextEdit
import openpyxl

now = datetime.now()


dtString = now.strftime("%H:%M")



path = r'C:\Users\Admin\PycharmProjects\face_detection\face_detection\images'
images = []
classNames = []
myList = os.listdir(path)
print(myList)
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])
print(classNames)

def markAttendance(name):
    attendance_file_path = 'Attendance.csv'

    # Create the file if it doesn't exist
    if not os.path.isfile(attendance_file_path):
        with open(attendance_file_path, 'w', newline='') as f:
            f.write('Name,Time,Date,Mark\n')

    # Open the file in append mode and write the attendance record
    with open(attendance_file_path, 'a', newline='') as f:
        f.write(f'{name},{dtString},{datetime.now().strftime("%Y-%m-%d")}\n')       
        

def findEncodings(images):
    encodeList =[]
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img)[0]
        encodeList.append(encode)
    return encodeList



def create_or_open_workbook():
    # Get the current date in the format "YYYY-MM-DD"
    
        current_date = datetime.now().strftime("%Y-%m-%d")
    #print(current_date)
        # Try to open an existing workbook with the current date as the sheet name
        try:
            workbook = openpyxl.load_workbook(f"{current_date}.xlsx")
        except FileNotFoundError:
           # If the workbook does not exist, create a new one
             workbook = openpyxl.Workbook()
             workbook.save(f"{current_date}.xlsx")

        # Return the active sheet
        sheet = workbook.active
        return workbook, sheet
    
def is_username_unique(sheet, username):
      for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            return False
        return True

def is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes):
    try:
        # Parse the current time and target time strings into datetime objects
        current_time = datetime.strptime(current_time_str, '%H:%M')
        target_time = datetime.strptime(target_time_str, '%H:%M')

        # Calculate the time difference between current time and target time
        time_difference = current_time - target_time

        # Convert the time difference to minutes
        lateness_minutes = time_difference.total_seconds() / 60

        # Check if the lateness is greater than or equal to the specified threshold
        if lateness_minutes >= lateness_threshold_minutes:
            return "Late"
        else:
            return "On Time"
    except ValueError:
        return "Invalid time format. Use 'HH:MM' format (e.g., '13:45')."

def main(user_input):
    workbook, sheet = create_or_open_workbook()

    # Add column headers if the sheet is empty
    if sheet.max_row == 1:
        sheet.append(["Username", "Current Date", "Current Time","Mark"])

       
    user_data = set()  # To store unique usernames

   # Check if the username is unique in the sheet
    if is_username_unique(sheet, user_input):
            user_data.add(user_input)
            current_time_str = dtString
            target_time_str = '18:00'
            lateness_threshold_minutes = 6
            result = is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes)
            print(result) 
            # Get the current date and time
            current_date = datetime.now().strftime("%Y-%m-%d")
            current_time = datetime.now().strftime("%H:%M:%S")

            # Append user input, current date, and current time to the next row
            sheet.append([user_input, current_date, current_time,result])
            
            workbook.save(f"{current_date}.xlsx")  # Save the workbook after each update
    else:
          print("Username already exists. Please enter a unique username.")

    



encodeListKnown = findEncodings(images)
print('Encoding Complete')

cap = cv2.VideoCapture(0)


while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        print(faceDis)
        matchIndex = np.argmin(faceDis)
        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
        else:
            name = 'Unknown'

        print(name)
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.rectangle(img, (x1, y2-35), (x2, y2), (0, 250, 0), cv2.FILLED)
        cv2.putText(img, name, (x1+6, y2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
        markAttendance(name)

        # Call create_or_open_workbook function
        create_or_open_workbook()
    
        
            
    cv2.imshow('FaceRecognition App', img)
    if cv2.waitKey(10) == 13:
        break
cap.release()
cv2.destroyAllWindows()

# Check if the person's attendance has been recorded in the last hour
if self.is_attendance_recorded_recently(person_name):
    log_message = f"Attendance already recorded for {person_name} within the last hour. Come back later."
    self.log_emitter.log_updated.emit(log_message)
else:
    # Record attendance
    date_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(self.attendance_file_path, 'a') as f:
        f.write(f"{person_name},{date_time_str}\n")
        f.close()
            self.log_emitter.log_updated.emit(f"Attendance recorded for {person_name}")
            # Update the last recorded time for the person
            self.last_attendance_recorded[person_name] = datetime.now()

